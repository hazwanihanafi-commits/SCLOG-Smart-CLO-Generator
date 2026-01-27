# ======================================================
# SCLOG — CLEAN FINAL BACKEND (EXCEL + LOGIC EXPLANATIONS)
# ======================================================

import os
import json
from io import BytesIO
from datetime import datetime
from flask import (
    Flask, render_template, jsonify, request,
    send_file
)
import pandas as pd
from openpyxl import Workbook, load_workbook

# ------------------------------------------------------
# App setup
# ------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    static_folder=os.path.join(BASE_DIR, "static"),
    template_folder=os.path.join(BASE_DIR, "templates")
)

WORKBOOK_PATH = os.path.join(BASE_DIR, "SCLOG (1).xlsx")
FRONT_JSON = os.path.join(app.static_folder, "data", "SCLOG_front.json")


# ------------------------------------------------------
# Safe JSON loader
# ------------------------------------------------------
def safe_load_json(path):
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

MAP = safe_load_json(FRONT_JSON)

DEFAULT_KEYS = {
    "IEGs": [], "PEOs": [], "PLOs": [],
    "IEGtoPEO": {}, "PEOtoPLO": {},
    "PLOstatements": {}, "PEOstatements": {},
    "PLOtoVBE": {}, "PLOIndicators": {},
    "SCmapping": {}
}
for k, v in DEFAULT_KEYS.items():
    MAP.setdefault(k, v)


# ------------------------------------------------------
# Excel helpers
# ------------------------------------------------------
def load_df(sheet_name):
    if not os.path.exists(WORKBOOK_PATH):
        return pd.DataFrame()
    try:
        return pd.read_excel(WORKBOOK_PATH, sheet_name=sheet_name, engine="openpyxl")
    except:
        return pd.DataFrame()


PROFILE_SHEET_MAP = {
    "health": "Mapping_health",
    "sc": "Mapping_sc",
    "eng": "Mapping_eng",
    "socs": "Mapping_socs",
    "edu": "Mapping_edu",
    "bus": "Mapping_bus",
    "arts": "Mapping_arts"
}

def get_mapping_sheet(profile):
    sheet = PROFILE_SHEET_MAP.get(profile, "Mapping")
    df = load_df(sheet)
    if df.empty:
        df = load_df("Mapping")
    return df


def get_plo_details(plo, profile="sc"):
    df = get_mapping_sheet(profile)
    if df.empty:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    col_plo = df.columns[0]
    mask = df[col_plo].astype(str).str.upper() == str(plo).upper()

    if not mask.any():
        return None

    row = df[mask].iloc[0]
    return {
        "SC_Code": row.get("SC Code", row.get("SCCode", "")),
        "SC_Desc": row.get("SC Description", row.get("SCDescription", "")),
        "VBE": row.get("VBE", ""),
        "Domain": row.get("Domain", "")
    }


# ------------------------------------------------------
# META (Criterion + Condition)
# ------------------------------------------------------
def get_meta_data(plo, bloom, profile="sc"):
    details = get_plo_details(plo, profile)
    if not details:
        return {}

    domain = (details.get("Domain") or "").lower()
    criterion = ""
    condition = ""

    df = load_df("Criterion")
    if not df.empty:
        df.columns = [c.strip() for c in df.columns]

        left = df.iloc[:,0].astype(str).str.lower()
        right = df.iloc[:,1].astype(str).str.lower()

        mask = (left == domain) & (right == str(bloom).lower())
        if mask.any():
            row = df[mask].iloc[0]
            if len(row) > 2:
                criterion = str(row.iloc[2])
            if len(row) > 3:
                condition = str(row.iloc[3])

    if not condition:
        defaults = {
            "cognitive": "interpreting tasks",
            "affective": "engaging with peers",
            "psychomotor": "performing skills"
        }
        condition = defaults.get(domain, "")

    connector = "by" if domain == "psychomotor" else "when"
    cond_final = f"{connector} {condition}"

    return {
        "sc_code": details["SC_Code"],
        "sc_desc": details["SC_Desc"],
        "vbe": details["VBE"],
        "domain": domain,
        "criterion": criterion,
        "condition": cond_final
    }


# ------------------------------------------------------
# Assessment / Evidence
# ------------------------------------------------------
def get_assessment(plo, bloom, domain, profile):
    b = bloom.lower().strip()
    d = domain.lower().strip()
    p = profile.strip().lower()

    # ===============================
    # COGNITIVE — BY PROFILE
    # ===============================
    cognitive = {

        "medical & health": {
            "remember": ["MCQ", "Quiz", "Recall questions"],
            "understand": ["Short answer", "Concept explanation"],
            "apply": ["Case-based discussion", "Short case", "Screening task"],
            "analyze": ["Case analysis", "Journal critique"],
            "analyse": ["Case analysis", "Journal critique"],
            "evaluate": ["Long case", "Viva Voce", "Clinical decision justification"],
            "create": ["Clinical management plan", "Health intervention proposal"]
        },

        "computer science & it": {
            "remember": ["MCQ", "Quiz"],
            "understand": ["Short answer", "Code explanation"],
            "apply": ["Programming assignment", "Coding exercise"],
            "analyze": ["Code analysis", "Debugging task"],
            "analyse": ["Code analysis", "Debugging task"],
            "evaluate": ["Code review", "System evaluation report"],
            "create": ["Software project", "Capstone project"]
        },

        "engineering & technology": {
            "remember": ["Test", "Quiz"],
            "understand": ["Technical explanation"],
            "apply": ["Problem-solving assignment", "Design exercise"],
            "analyze": ["System analysis", "Technical report"],
            "analyse": ["System analysis", "Technical report"],
            "evaluate": ["Design evaluation", "Oral presentation"],
            "create": ["Design project", "Capstone project"]
        },

        "social sciences": {
            "remember": ["Test", "Reading quiz"],
            "understand": ["Essay", "Discussion"],
            "apply": ["Case study", "Fieldwork report"],
            "analyze": ["Thematic analysis", "Policy analysis"],
            "analyse": ["Thematic analysis", "Policy analysis"],
            "evaluate": ["Critical review", "Oral presentation"],
            "create": ["Research project", "Policy proposal"]
        },

        "education": {
            "remember": ["Test", "Quiz"],
            "understand": ["Essay", "Reflection"],
            "apply": ["Lesson plan", "Microteaching"],
            "analyze": ["Teaching reflection report"],
            "analyse": ["Teaching reflection report"],
            "evaluate": ["Teaching evaluation", "Portfolio review"],
            "create": ["Curriculum design project", "Action research"]
        },

        "business & management": {
            "remember": ["Test", "Quiz"],
            "understand": ["Essay", "Case discussion"],
            "apply": ["Business case study", "Problem-solving assignment"],
            "analyze": ["Financial analysis", "Market analysis"],
            "analyse": ["Financial analysis", "Market analysis"],
            "evaluate": ["Strategy evaluation", "Oral presentation"],
            "create": ["Business plan", "Consultancy project"]
        },

        "arts & humanities": {
            "remember": ["Quiz", "Visual identification"],
            "understand": ["Essay", "Artwork interpretation"],
            "apply": ["Studio exercise", "Creative task"],
            "analyze": ["Artwork analysis", "Critical review"],
            "analyse": ["Artwork analysis", "Critical review"],
            "evaluate": ["Portfolio critique", "Oral presentation"],
            "create": ["Creative project", "Final portfolio"]
        }
    }

    # ===============================
    # AFFECTIVE — BY PROFILE
    # ===============================
    affective = {

        "medical & health": {
            "receive": ["Professional awareness reflection"],
            "respond": ["Ward / clinical participation"],
            "value": ["Ethics & patient safety reflection"],
            "organization": ["Interprofessional teamwork portfolio"],
            "characterization": ["Clinical professionalism assessment"]
        },

        "computer science & it": {
            "receive": ["Learning reflection"],
            "respond": ["Participation in technical discussions"],
            "value": ["Ethics in computing essay"],
            "organization": ["Team-based software project portfolio"],
            "characterization": ["Professional conduct in computing"]
        },

        "engineering & technology": {
            "receive": ["Safety awareness reflection"],
            "respond": ["Lab participation"],
            "value": ["Engineering ethics reflection"],
            "organization": ["Project team portfolio"],
            "characterization": ["Professional engineering behaviour"]
        },

        "social sciences": {
            "receive": ["Social awareness reflection"],
            "respond": ["Seminar participation"],
            "value": ["Ethical reasoning essay"],
            "organization": ["Group research portfolio"],
            "characterization": ["Professional social conduct"]
        },

        "education": {
            "receive": ["Teaching values reflection"],
            "respond": ["Classroom participation"],
            "value": ["Ethics in education essay"],
            "organization": ["Teaching portfolio"],
            "characterization": ["Teacher professionalism assessment"]
        },

        "business & management": {
            "receive": ["Business awareness reflection"],
            "respond": ["Case discussion participation"],
            "value": ["Business ethics essay"],
            "organization": ["Team consultancy portfolio"],
            "characterization": ["Professional business conduct"]
        },

        "arts & humanities": {
            "receive": ["Creative awareness reflection"],
            "respond": ["Studio participation"],
            "value": ["Artistic values reflection"],
            "organization": ["Creative portfolio"],
            "characterization": ["Professional artistic practice"]
        }
    }

    # ===============================
    # PSYCHOMOTOR — BY PROFILE
    # ===============================
    psychomotor = {

        "medical & health": {
            "perception": ["Recognition of clinical signs"],
            "set": ["Clinical preparation checklist"],
            "guided response": ["Supervised clinical task"],
            "mechanism": ["Clinical skills test", "OSCE"],
            "complex overt response": ["OSCE", "Clinical simulation"],
            "adaptation": ["Management of complex patients"],
            "origination": ["Independent patient management"]
        },

        "computer science & it": {
            "perception": ["Recognition of system requirements"],
            "set": ["Development environment setup"],
            "guided response": ["Guided coding task"],
            "mechanism": ["Hands-on coding test"],
            "complex overt response": ["System simulation"],
            "adaptation": ["Code optimisation task"],
            "origination": ["Independent software development"]
        },

        "engineering & technology": {
            "perception": ["Identification of system components"],
            "set": ["Lab setup checklist"],
            "guided response": ["Guided laboratory task"],
            "mechanism": ["Laboratory practical"],
            "complex overt response": ["Integrated lab assessment"],
            "adaptation": ["System troubleshooting"],
            "origination": ["Independent engineering task"]
        },

        "social sciences": {
            "perception": ["Observation of social phenomena"],
            "set": ["Fieldwork preparation"],
            "guided response": ["Guided data collection"],
            "mechanism": ["Fieldwork practical"],
            "complex overt response": ["Community-based simulation"],
            "adaptation": ["Contextual analysis task"],
            "origination": ["Independent field study"]
        },

        "education": {
            "perception": ["Classroom observation"],
            "set": ["Lesson preparation"],
            "guided response": ["Guided teaching practice"],
            "mechanism": ["Microteaching practical"],
            "complex overt response": ["Teaching simulation"],
            "adaptation": ["Adaptive teaching task"],
            "origination": ["Independent teaching session"]
        },

        "business & management": {
            "perception": ["Observation of business processes"],
            "set": ["Business case preparation"],
            "guided response": ["Guided business simulation"],
            "mechanism": ["Business skills practical"],
            "complex overt response": ["Management simulation"],
            "adaptation": ["Strategic adjustment task"],
            "origination": ["Independent consultancy task"]
        },

        "arts & humanities": {
            "perception": ["Observation of artistic techniques"],
            "set": ["Studio preparation"],
            "guided response": ["Guided studio task"],
            "mechanism": ["Studio practical"],
            "complex overt response": ["Performance / exhibition simulation"],
            "adaptation": ["Creative adaptation task"],
            "origination": ["Independent creative production"]
        }
    }

    if d == "cognitive":
        return cognitive.get(p, {}).get(b, [])
    if d == "affective":
        return affective.get(p, {}).get(b, [])
    if d == "psychomotor":
        return psychomotor.get(p, {}).get(b, [])

    return []

def get_evidence_for(assessment):
    a = assessment.lower().strip()

    mapping = {

        # TEST / QUIZ
        "mcq": ["Score report"],
        "quiz": ["Quiz score"],
        "test": ["Test score report"],
        "recall": ["Marked answer script"],

        # WRITTEN / ESSAY
        "short answer": ["Marked answer script"],
        "essay": ["Written essay"],
        "concept explanation": ["Written explanation"],
        "code explanation": ["Annotated code explanation"],
        "technical explanation": ["Written technical explanation"],

        # CASE / DISCUSSION
        "case-based discussion": ["CbD record", "Supervisor feedback"],
        "short case": ["Short case assessment form"],
        "long case": ["Long case report", "Examiner evaluation form"],
        "case study": ["Case study report"],
        "case analysis": ["Case analysis worksheet"],
        "discussion": ["Discussion participation record"],
        "journal critique": ["Journal critique report"],

        # CLINICAL / PRACTICAL
        "screening": ["Screening checklist"],
        "skills test": ["Skills checklist"],
        "osce": ["OSCE score sheet"],
        "simulation": ["Simulation checklist"],
        "observation": ["Observation checklist"],
        "guided task": ["Supervisor observation form"],

        # PROGRAMMING / IT
        "programming assignment": ["Source code submission", "Grading rubric"],
        "coding exercise": ["Code submission"],
        "debugging": ["Debugging report"],
        "code analysis": ["Code review report"],
        "code review": ["Code review rubric"],

        # ENGINEERING / DESIGN
        "design exercise": ["Design documentation"],
        "design project": ["Project report", "Design artefact"],
        "system analysis": ["System analysis report"],
        "technical report": ["Technical report"],

        # EDUCATION
        "lesson plan": ["Lesson plan document"],
        "microteaching": ["Teaching observation rubric"],
        "teaching evaluation": ["Teaching evaluation form"],
        "portfolio review": ["Portfolio evidence"],

        # BUSINESS / SOCIAL SCIENCE
        "financial analysis": ["Financial analysis report"],
        "market analysis": ["Market analysis report"],
        "policy analysis": ["Policy analysis report"],
        "fieldwork": ["Fieldwork report"],
        "consultancy": ["Consultancy report"],

        # PROJECT / RESEARCH / CREATIVE
        "project": ["Project documentation"],
        "capstone": ["Capstone project report"],
        "research": ["Research report"],
        "proposal": ["Proposal document"],
        "business plan": ["Business plan document"],
        "creative project": ["Creative artefact", "Project reflection"],
        "portfolio": ["Portfolio evidence"],

        # AFFECTIVE / PROFESSIONAL
        "reflection": ["Reflection journal"],
        "participation": ["Participation record"],
        "peer feedback": ["Peer feedback form"],
        "professional": ["Professional behaviour evaluation"],
        "ethics": ["Ethics reflection"],
        "presentation": ["Presentation rubric"]
    }

    evidence = []
    for key, items in mapping.items():
        if key in a:
            evidence.extend(items)

    evidence = list(dict.fromkeys(evidence))
    return evidence if evidence else ["Assessment evidence"]

# ------------------------------------------------------
# CONTENT suggestions
# ------------------------------------------------------
CONTENT_SUGGESTIONS = {
    "Computer Science": [
        "design software modules", "analyze data structures", "build machine learning models"
    ],
    "Medical & Health": [
        "interpret ECG", "analyze rehabilitation progress", "perform screenings"
    ],
    "Engineering": [
        "apply thermodynamics", "analyze structural loads"
    ],
    "Education": [
        "design lesson plans", "evaluate learning outcomes"
    ]
}

@app.route("/api/content/<field>")
def api_content(field):
    for k in CONTENT_SUGGESTIONS:
        if k.lower() == field.lower():
            return jsonify(CONTENT_SUGGESTIONS[k])
    return jsonify([])


# ------------------------------------------------------
# MAPPING endpoints (IEG → PEO → PLO)
# ------------------------------------------------------
@app.route("/api/mapping")
def api_mapping():
    return jsonify(MAP)

@app.route("/api/get_peos/<ieg>")
def api_get_peos(ieg):
    return jsonify(MAP["IEGtoPEO"].get(ieg, []))

@app.route("/api/get_plos/<peo>")
def api_get_plos(peo):
    return jsonify(MAP["PEOtoPLO"].get(peo, []))


# ------------------------------------------------------
# LOGIC explanations
# ------------------------------------------------------
@app.route("/api/logic/ieg_peo/<ieg>")
def logic_ieg_peo(ieg):
    explanation = {
        "IEG1": "IEG1 focuses on knowledge & critical thinking. PEO1 operationalises these outcomes.",
        "IEG2": "IEG2 emphasises ethics & professionalism. PEO2 aligns with these values.",
        "IEG3": "IEG3 promotes socio-entrepreneurship. PEO3 guides this development.",
        "IEG4": "IEG4 strengthens communication. PEO4 builds communication competence.",
        "IEG5": "IEG5 focuses on leadership & lifelong learning. PEO5 supports these traits."
    }
    return explanation.get(ieg, "No logic found."), 200, {"Content-Type": "text/plain"}


@app.route("/api/logic/peo_plo/<peo>/<plo>")
def logic_peo_plo(peo, plo):
    logic_map = {
        "PEO1": {
            "PLO1":"Knowledge → foundation for PEO1",
            "PLO2":"Critical thinking → supports PEO1",
            "PLO3":"Analysis ability → supports PEO1",
            "PLO6":"Real-world application → supports PEO1",
            "PLO7":"Problem-solving → supports PEO1"
        },
        "PEO2": {"PLO11":"Ethics & professionalism → supports PEO2"},
        "PEO3": {"PLO9":"Sustainability → supports PEO3","PLO10":"Societal wellbeing → supports PEO3"},
        "PEO4": {"PLO5":"Communication skills → supports PEO4"},
        "PEO5": {"PLO4":"Teamwork → supports PEO5","PLO8":"Leadership → supports PEO5","PLO9":"Global challenges → supports PEO5"}
    }
    return logic_map.get(peo, {}).get(plo, "No logic available."), 200, {"Content-Type":"text/plain"}


# ------------------------------------------------------
# BLOOM & VERB endpoints (Excel)
# ------------------------------------------------------
@app.route("/api/get_blooms/<plo>")
def api_get_blooms(plo):
    profile = request.args.get("profile", "sc").lower()

    details = get_plo_details(plo, profile)
    if not details:
        return jsonify([])

    domain = str(details.get("Domain", "")).strip().lower()

    sheet_map = {
        "cognitive": "Bloom_Cognitive",
        "affective": "Bloom_Affective",
        "psychomotor": "Bloom_Psychomotor"
    }

    sheet = sheet_map.get(domain)
    if not sheet:
        return jsonify([])

    df = load_df(sheet)
    if df.empty:
        return jsonify([])

    # ✅ BACA MENGGUNAKAN NAMA COLUMN SEBENAR EXCEL
    if "Bloom Level" not in df.columns:
        return jsonify([])

    blooms = (
        df["Bloom Level"]
        .dropna()
        .astype(str)
        .str.strip()
        .tolist()
    )

    return jsonify(blooms)




# ------------------------------------------------------
# GET VERBS (BY BLOOM ONLY) — NEW
# ------------------------------------------------------
@app.route("/api/get_verbs/<bloom>")
def api_get_verbs_by_bloom(bloom):
    bloom_key = bloom.strip().lower()

    # Try all Bloom sheets (Bloom taxonomy is domain-based)
    for sheet in ["Bloom_Cognitive", "Bloom_Affective", "Bloom_Psychomotor"]:
        df = load_df(sheet)
        if df.empty:
            continue

        if "Bloom Level" not in df.columns:
            continue

        mask = df["Bloom Level"].astype(str).str.lower() == bloom_key
        if mask.any():
            # Column 2 = verbs (based on your Excel)
            raw = df.loc[mask].iloc[0, 1]
            verbs = [v.strip() for v in str(raw).split(",") if v.strip()]
            return jsonify(verbs)

    return jsonify([])


# ------------------------------------------------------
# META endpoint
# ------------------------------------------------------
@app.route("/api/get_meta/<plo>/<bloom>")
def api_get_meta(plo, bloom):
    profile = request.args.get("profile","sc").lower()
    return jsonify(get_meta_data(plo, bloom, profile))


# ------------------------------------------------------
# STATEMENT endpoint
# ------------------------------------------------------
@app.route("/api/get_statement/<level>/<stype>/<code>")
def api_get_statement(level, stype, code):
    if stype == "PEO":
        return jsonify(MAP["PEOstatements"].get(level, {}).get(code, ""))
    if stype == "PLO":
        return jsonify(MAP["PLOstatements"].get(level, {}).get(code, ""))
    return jsonify("")

# ------------------------------------------------------
# GLOBAL STATE
# ------------------------------------------------------
LAST_CLO = {}


# ------------------------------------------------------
# GENERATE CLO
# ------------------------------------------------------

@app.route("/generate", methods=["POST"])
def generate():
    global LAST_CLO

    # ----------------------------------
    # NORMALISE PROFILE FROM UI
    # ----------------------------------
    PROFILE_ALIAS = {
        "sc": "computer science & it",
        "health": "medical & health",
        "eng": "engineering & technology",
        "socs": "social sciences",
        "edu": "education",
        "bus": "business & management",
        "arts": "arts & humanities"
    }

    raw_profile = request.form.get("profile", "medical & health").strip().lower()
    profile = PROFILE_ALIAS.get(raw_profile, raw_profile)

    # ----------------------------------
    # CONTINUE NORMAL FLOW
    # ----------------------------------
    plo = request.form.get("plo", "")
    bloom = request.form.get("bloom", "")
    verb = request.form.get("verb", "")
    if not verb:
        verb = bloom.lower()   # fallback: remember, analyze, etc.
    content = request.form.get("content", "")
    level = request.form.get("level", "Degree")
    programme_name = request.form.get("programmeName", "")
    ieg_input = request.form.get("ieg", "").strip()
    course_name = request.form.get("courseName", "")
    peo_statement = request.form.get("peo_statement", "").strip()
    plo_indicator = request.form.get("plo_indicator", "").strip()

    # ✅ REQUIRED FIELD CHECK
    if not plo or not bloom or not verb or not content:
        return jsonify({"error": "Missing required fields"}), 400
    
    details = get_plo_details(plo, profile)
    if not details:
        return jsonify({"error": "Invalid PLO"}), 400

    meta = get_meta_data(plo, bloom, profile)

    domain = details["Domain"].lower()
    sc_desc = details["SC_Desc"]
    vbe = details["VBE"]

    # Clean verb duplication
    words = content.strip().split()
    if words and words[0].lower() == verb.lower():
        content = " ".join(words[1:])

    connector = "when" if domain != "psychomotor" else "by"
    condition_clean = meta["condition"].replace("when ", "").replace("by ", "")

    clo = (
        f"{verb.lower()} {content} using {sc_desc.lower()} "
        f"{connector} {condition_clean} guided by {vbe.lower()}."
    ).capitalize()

    variants = {
        "Standard": clo,
        "Critical Thinking": clo.replace("using", "critically using"),
        "Short": f"{verb.capitalize()} {content}."
    }

    peo = next((p for p, plos in MAP["PEOtoPLO"].items() if plo in plos), None)

    if ieg_input:
        ieg = ieg_input
    else:
        ieg = next((i for i, peos in MAP["IEGtoPEO"].items() if peo in peos), "Paste IEG")

    assessments = get_assessment(plo, bloom, domain, profile)
    evidence = {a: get_evidence_for(a) for a in assessments}


    LAST_CLO = {
    # ======================
    # PROGRAMME CONTEXT
    # ======================
    "programme_name": programme_name,
    "course_name": course_name,
    "ieg": ieg,

    # ======================
    # PEO
    # ======================
    "peo": peo,
     "peo_statement": (
    peo_statement
    if peo_statement
    else MAP.get("PEOstatements", {}).get(peo, "")
),

    # ======================
    # PLO
    # ======================
    "plo": plo,
    "plo_statement": MAP["PLOstatements"].get(level, {}).get(
        plo, "Full MQF-aligned PLO"
    ),
    "plo_indicator": (
    plo_indicator
    if plo_indicator
    else "; ".join(
        MAP.get("PLOindicators", {})
        .get(level, {})
        .get(plo, [])
    )
),

        

    # ======================
    # CLO
    # ======================
    "clo": clo,
    "clo_indicator": "≥60% achievement",
    "variants": variants,

    # ======================
    # ASSESSMENT
    # ======================
    "assessments": assessments,
    "evidence": evidence,

    # ======================
    # MQF / VBE
    # ======================
    "sc_code": details["SC_Code"],
    "sc_desc": details["SC_Desc"],
    "domain": details["Domain"],
    "condition": meta["condition"],
    "criterion": meta["criterion"],
    "vbe": details["VBE"]
}

    return jsonify(LAST_CLO)


# ------------------------------------------------------
# DOWNLOADS
# ------------------------------------------------------
@app.route("/download")
def download_clo():
    if not LAST_CLO:
        return "No CLO generated", 400

    wb = Workbook()
    ws = wb.active
    ws.title = "CLO"

    ws.append(["Field", "Value"])

    for key, val in LAST_CLO.items():
        if isinstance(val, dict):
            val = json.dumps(val, ensure_ascii=False)
        elif isinstance(val, list):
            val = "; ".join(str(x) for x in val)
        ws.append([key, val])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        as_attachment=True,
        download_name=f"CLO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/download_rubric")
def download_rubric():
    if not LAST_CLO:
        return "Generate CLO first", 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Rubric"

    ws.append(["Component","Description"])
    ws.append(["Indicator", f"Ability to {LAST_CLO['clo']}"])
    ws.append(["Excellent","Performs at excellent level"])
    ws.append(["Good","Performs well"])
    ws.append(["Satisfactory","Meets minimum level"])
    ws.append(["Poor","Below expected"])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"Rubric_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        out, as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ------------------------------------------------------
# UI
# ------------------------------------------------------
@app.route("/")
def landing():
    return render_template("landing.html")

@app.route("/app")
def workflow():
    return render_template("index.html")

@app.route("/clo-only")
def clo_only_page():
    return render_template("clo_only.html")

from server import clo_only as clo_only_bp
app.register_blueprint(clo_only_bp)


# ------------------------------------------------------
# RUN
# ------------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")




































